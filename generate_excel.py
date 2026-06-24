"""Generates EasyTravel_Sample_Transcripts.xlsx with 3 sheets:
  1. Sample Transcripts — 10 rows, Scenario | Full_Transcript | Expected_Summary
  2. Evaluation Sheet   — 10 rows, Scenario | AI_Generated_Summary | Your_Evaluation | Notes
  3. Answer Key         — hidden, Scenario | Correct_Summary | Known_Flaw_Type | What_Was_Wrong
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent / "sample_data"
OUTPUT_FILE = OUTPUT_DIR / "EasyTravel_Sample_Transcripts.xlsx"


def _t(*turns: tuple[str, str]) -> str:
    return "\n".join(f"{speaker}: {text}" for speaker, text in turns)


TRANSCRIPTS = [
    {
        "scenario": "Flight cancellation refund — resolved successfully",
        "transcript": _t(
            ("Agent", "Thank you for calling EasyTravel, this is Priya. How may I help you today?"),
            ("Customer", "Hi Priya, this is Rohan Mehta. I need to cancel my flight ET-4521 from Bangalore to Delhi on the 15th of July."),
            ("Agent", "I am sorry to hear that, Rohan. May I know the reason for cancellation? It helps us process the right type of refund."),
            ("Customer", "I have a medical issue, I cannot travel. I have a doctor's certificate if you need it."),
            ("Agent", "I understand. For medical cancellations we waive the cancellation fee if you can email the certificate within 7 days."),
            ("Customer", "That's great. What about the refund?"),
            ("Agent", "Your booking total was 4,500 rupees. The full amount of 4,500 will be refunded to the original payment method within 7 business days."),
            ("Customer", "Perfect. Is there anything else I need to do?"),
            ("Agent", "Please send the medical certificate to refunds@easytravel.example with subject 'ET-4521 medical'. I will note the cancellation and waiver on the booking now."),
            ("Customer", "Done, I will email it today."),
            ("Agent", "Excellent. I am processing the cancellation now. You will receive a cancellation confirmation email at rohan.m@example.com within the next 10 minutes."),
            ("Customer", "Thank you so much. That was very smooth."),
            ("Agent", "You're welcome. I am also sending a follow-up email confirming the refund timeline and the certificate instructions."),
            ("Customer", "Great, thanks again Priya. Bye."),
            ("Agent", "Take care, Rohan. Have a speedy recovery."),
        ),
        "expected": (
            "Customer Rohan Mehta requested cancellation of flight ET-4521 (BLR-DEL, 15 Jul) "
            "due to a medical reason. Agent Priya waived the cancellation fee on medical grounds "
            "and confirmed a full refund of Rs. 4,500 within 7 business days. "
            "Customer agreed to email the medical certificate within 7 days. Resolved."
        ),
        "correct": (
            "Issue: Cancellation of flight ET-4521 due to medical reason. "
            "Resolution: Resolved — cancellation processed, fee waived, Rs. 4,500 refund within 7 business days. "
            "Action items: Send cancellation confirmation email; customer to email medical certificate to refunds@easytravel.example within 7 days. "
            "Sentiment: Satisfied. "
            "Key details: Booking ET-4521, BLR-DEL, 15-Jul, refund Rs. 4,500, agent Priya, customer Rohan Mehta."
        ),
        "flaw": "wrong_resolution",
        "what_wrong": "A flawed summary may mark this as 'Escalated' or 'Pending' even though the refund was fully processed.",
    },
    {
        "scenario": "Seat upgrade — partial resolution, callback promised",
        "transcript": _t(
            ("Agent", "Good afternoon, EasyTravel support, Rahul speaking."),
            ("Customer", "Hello Rahul, this is Anita Sharma. I want a seat upgrade on my flight ET-8810 next Friday."),
            ("Agent", "Let me check that booking for you. One moment please. ... Yes I have it — Mumbai to Singapore, departing 12 July, currently economy 23A."),
            ("Customer", "Yes that's the one. I'd like to upgrade to business class."),
            ("Agent", "Business class availability shows 2 seats remaining. The upgrade cost would be 18,000 rupees over your current fare."),
            ("Customer", "That's higher than I expected. Are there any options for an upgrade with miles?"),
            ("Agent", "Let me check your loyalty account... You have 32,000 EasyMiles. A business upgrade for this sector needs 40,000 miles plus a 4,500 rupee tax."),
            ("Customer", "So I am 8,000 miles short."),
            ("Agent", "Yes. However we sometimes do mileage waivers for our Gold members on a case-by-case basis. You are a Gold member."),
            ("Customer", "Can you check if I can get that waiver?"),
            ("Agent", "I will need to send a request to our loyalty desk. They typically respond within 24 hours."),
            ("Customer", "Alright please raise the request."),
            ("Agent", "Request raised, reference UPG-66422. I will personally call you back tomorrow before 6 PM with the loyalty team's decision."),
            ("Customer", "Tomorrow before 6 PM. Got it. My number is correct in the system right?"),
            ("Agent", "Yes I have +91 98XXX 22113. I will call you on that. Anything else for now?"),
            ("Customer", "No that's all, thanks Rahul."),
            ("Agent", "Thank you for being a Gold member, Anita. Talk to you tomorrow."),
        ),
        "expected": (
            "Customer Anita Sharma (Gold member) requested a business class upgrade on ET-8810 "
            "(BOM-SIN, 12 Jul). Cost would be Rs. 18,000 or 40,000 miles + Rs. 4,500 tax. "
            "Customer is 8,000 miles short. Agent Rahul raised mileage-waiver request UPG-66422 "
            "with loyalty desk and committed to a callback before 6 PM the next day. Pending."
        ),
        "correct": (
            "Issue: Business class upgrade request on ET-8810 (BOM-SIN, 12 Jul), customer 8,000 miles short. "
            "Resolution: Pending — mileage waiver request UPG-66422 raised with loyalty desk, callback scheduled. "
            "Action items: Agent to call back before 6 PM the next day with loyalty team's decision on miles waiver. "
            "Sentiment: Neutral (cooperative). "
            "Key details: Booking ET-8810, BOM-SIN, 12-Jul, 32k miles, request UPG-66422, agent Rahul, customer Anita Sharma (Gold)."
        ),
        "flaw": "missed_action_item",
        "what_wrong": "A flawed summary may omit the callback commitment for the next day before 6 PM.",
    },
    {
        "scenario": "Lost baggage claim — escalated to supervisor",
        "transcript": _t(
            ("Agent", "EasyTravel support, this is Ananya. How can I help?"),
            ("Customer", "Hi, I am Vikram Iyer. My checked-in baggage did not arrive in Chennai. My flight was ET-2104 from Dubai yesterday."),
            ("Agent", "I am very sorry to hear that, Vikram. Did you file a Property Irregularity Report at the airport?"),
            ("Customer", "Yes I did, PIR number MAA-9981. The counter staff said they will trace it but no one has called me."),
            ("Agent", "Let me look up the PIR. ... I can see the trace is in progress but no match yet."),
            ("Customer", "It's been more than 30 hours. I have important medicines in that bag."),
            ("Agent", "I completely understand. Let me see what immediate assistance I can offer. As per policy you are entitled to an interim allowance of 5,000 rupees for essentials."),
            ("Customer", "How do I claim that?"),
            ("Agent", "I will initiate it now — it will reflect as a UPI transfer to your registered mobile within 24 hours."),
            ("Customer", "Okay. But what about the bag itself? I really need it back."),
            ("Agent", "I want to escalate this to our supervisor in the baggage tracing team for priority handling, given the medicines."),
            ("Customer", "Yes please escalate."),
            ("Agent", "I am transferring this case to Senior Supervisor Manish Khanna. He will take ownership and contact you within 4 hours."),
            ("Customer", "Will I get an escalation reference?"),
            ("Agent", "Yes — escalation ID is ESC-44217. Manish will use this when he calls."),
            ("Customer", "Alright. Thank you for your help Ananya."),
            ("Agent", "Of course. I am noting the medical urgency on the file. You will hear from Manish within 4 hours."),
        ),
        "expected": (
            "Customer Vikram Iyer reported missing checked-in baggage from flight ET-2104 (DXB-MAA), "
            "PIR MAA-9981. Bag contains important medicines. Agent Ananya initiated Rs. 5,000 interim "
            "allowance and escalated to Senior Supervisor Manish Khanna (ESC-44217) with a 4-hour "
            "callback SLA. Escalated."
        ),
        "correct": (
            "Issue: Missing checked-in baggage on flight ET-2104 (DXB-MAA), PIR MAA-9981, contains medicines. "
            "Resolution: Escalated to Senior Supervisor Manish Khanna under escalation ESC-44217. "
            "Action items: Rs. 5,000 interim allowance UPI transfer within 24 hours; supervisor callback within 4 hours. "
            "Sentiment: Frustrated (initially) becoming Neutral after escalation. "
            "Key details: ET-2104, DXB-MAA, PIR MAA-9981, ESC-44217, agent Ananya, customer Vikram Iyer."
        ),
        "flaw": "wrong_sentiment",
        "what_wrong": "A flawed summary may label the customer as 'Satisfied' although they expressed clear frustration about the medicines and delayed callback.",
    },
    {
        "scenario": "Travel insurance claim query — frustrated, unresolved",
        "transcript": _t(
            ("Agent", "EasyTravel customer service, this is Karan. How may I help?"),
            ("Customer", "Hello Karan, this is Meena Joseph. I have been calling for three days about my travel insurance claim and no one is helping me."),
            ("Agent", "I am very sorry for the experience. Please share the claim reference."),
            ("Customer", "It is TI-CLM-77819. My husband fell ill in Bali and we had hospitalisation expenses of around 1,80,000 rupees."),
            ("Agent", "Let me check that case... I see the claim is logged but is awaiting documents."),
            ("Customer", "I have already emailed all documents twice — discharge summary, bills, prescription."),
            ("Agent", "Let me verify whether the documents are received in our claims portal... I can see the discharge summary and bills, but the system is showing 'prescription pending'."),
            ("Customer", "That cannot be right. I sent the prescription twice. Please check the email I sent on the 4th."),
            ("Agent", "I understand. The insurance is handled by our partner UrTravelGuard. I do not have direct access to their inbox."),
            ("Customer", "So what do I do? I have been told this every time I call."),
            ("Agent", "I will personally raise an internal ticket with UrTravelGuard requesting an immediate review of your emails and a status update."),
            ("Customer", "Honestly Karan, I have lost faith. I want a complaint registered."),
            ("Agent", "I understand. I will register a formal complaint with reference COMP-2231 and a parallel follow-up on the claim itself."),
            ("Customer", "What is the next step? I need actual progress, not another follow-up."),
            ("Agent", "I will email you today by 7 PM with confirmation that the complaint is registered and that UrTravelGuard has acknowledged the prescription document."),
            ("Customer", "Okay. If I don't hear from you by 7 PM I will call again, and please make sure your manager is aware of this."),
            ("Agent", "Absolutely. I am flagging this with my floor manager Sanjay Verma right now. I am sorry for the frustration."),
        ),
        "expected": (
            "Customer Meena Joseph has been chasing travel insurance claim TI-CLM-77819 for 3 days "
            "for Rs. 1,80,000 hospitalisation in Bali. System shows 'prescription pending' although "
            "customer has emailed it twice. Agent Karan raised complaint COMP-2231 and is flagging "
            "to floor manager Sanjay Verma. Promised an email update by 7 PM. Unresolved/Pending."
        ),
        "correct": (
            "Issue: Travel insurance claim TI-CLM-77819 stuck; partner UrTravelGuard claims prescription is missing although customer emailed it twice. "
            "Resolution: Pending — complaint COMP-2231 raised, flagged to floor manager Sanjay Verma. "
            "Action items: Karan to email confirmation by 7 PM that complaint is registered and prescription is acknowledged; UrTravelGuard to review and respond. "
            "Sentiment: Frustrated. "
            "Key details: TI-CLM-77819, Rs. 1,80,000, Bali hospitalisation, COMP-2231, agent Karan, customer Meena Joseph."
        ),
        "flaw": "fabricated_detail",
        "what_wrong": "A flawed summary may invent a refund amount or a fictitious claim reference that was not actually in the transcript.",
    },
    {
        "scenario": "Loyalty points redemption — resolved with workaround",
        "transcript": _t(
            ("Agent", "Hello, EasyTravel loyalty support, this is Sneha."),
            ("Customer", "Hi Sneha, I am Arjun Kapoor. I have 65,000 EasyMiles and I want to redeem them for a return ticket from Delhi to Bangkok."),
            ("Agent", "Wonderful, let me check availability. Which dates are you looking at?"),
            ("Customer", "Outbound 20 August, return 27 August."),
            ("Agent", "I can see an award seat outbound on 20 August but the return on 27 August is fully booked in the award inventory."),
            ("Customer", "Oh no, I really need those dates."),
            ("Agent", "One option — we can lock the outbound now using miles, and you can either change the return date or pay cash for the return. The cash for 27 August return is 14,800 rupees."),
            ("Customer", "Hmm. Is there any chance of an award opening up?"),
            ("Agent", "Award inventory does change, but I cannot guarantee it. The risk is the cash fare also keeps going up if you wait."),
            ("Customer", "Okay let's do this — book the outbound with miles, and let me pay cash for the return."),
            ("Agent", "Excellent decision. So 35,000 miles for the outbound + taxes 1,800 rupees, plus 14,800 cash for the return."),
            ("Customer", "Total cash today?"),
            ("Agent", "Rs. 16,600 — that's taxes on the miles ticket plus the return fare."),
            ("Customer", "Got it. Process it please."),
            ("Agent", "Payment received. PNRs: outbound DEL-BKK is PNR HG7XQ2, return BKK-DEL is PNR PL9MR4."),
            ("Customer", "Thanks. Will I get an email confirmation?"),
            ("Agent", "Yes, both confirmations are emailed to arjun.k@example.com. Travel safely on the 20th."),
        ),
        "expected": (
            "Customer Arjun Kapoor wanted to redeem 65,000 EasyMiles for return DEL-BKK 20-27 Aug. "
            "Return award was unavailable, so agent Sneha booked outbound on miles (35,000 + Rs. 1,800 tax) "
            "and cash return at Rs. 14,800. Total paid Rs. 16,600. PNRs: HG7XQ2 (out), PL9MR4 (return). Resolved with workaround."
        ),
        "correct": (
            "Issue: Loyalty redemption for DEL-BKK 20-27 Aug; return award seat unavailable. "
            "Resolution: Resolved with workaround — outbound on miles (35,000 + Rs. 1,800 tax), return paid in cash (Rs. 14,800). Total paid Rs. 16,600. "
            "Action items: Send email confirmations for both PNRs to customer. "
            "Sentiment: Satisfied. "
            "Key details: PNR HG7XQ2 (outbound), PNR PL9MR4 (return), 65,000 EasyMiles balance, agent Sneha, customer Arjun Kapoor."
        ),
        "flaw": "wrong_number",
        "what_wrong": "A flawed summary may shift the total amount or the miles required by a small amount.",
    },
    {
        "scenario": "Hotel booking amendment — simple resolution",
        "transcript": _t(
            ("Agent", "EasyTravel hotels, you have Divya speaking."),
            ("Customer", "Hi Divya, I am Sameer Khan. I want to change the check-in date on my Goa hotel booking."),
            ("Agent", "Sure, may I have the booking reference?"),
            ("Customer", "ETH-GOA-55821, currently for 10 September check-in."),
            ("Agent", "I see it — Hotel Solora, 10 to 13 September, deluxe room. What date would you like?"),
            ("Customer", "11 September to 14 September."),
            ("Agent", "Let me check availability... Yes the room type is available for the new dates."),
            ("Customer", "Are there any change fees?"),
            ("Agent", "For changes made more than 7 days before check-in there is no change fee, only any fare difference."),
            ("Customer", "What is the fare difference?"),
            ("Agent", "The new dates are actually 350 rupees lower in total. We will refund the 350 to your original card."),
            ("Customer", "Lovely. Please go ahead."),
            ("Agent", "Done. Your booking is now Hotel Solora, 11 to 14 September, deluxe room. New booking confirmation has been emailed to sameer.k@example.com."),
            ("Customer", "Perfect, thanks!"),
            ("Agent", "You're welcome, enjoy your trip to Goa."),
        ),
        "expected": (
            "Customer Sameer Khan amended hotel booking ETH-GOA-55821 (Hotel Solora, deluxe room) "
            "from 10-13 Sep to 11-14 Sep. No change fee (more than 7 days prior); fare difference of "
            "Rs. 350 to be refunded. New confirmation emailed. Resolved."
        ),
        "correct": (
            "Issue: Date change on hotel booking ETH-GOA-55821 (Hotel Solora, Goa) from 10-13 Sep to 11-14 Sep. "
            "Resolution: Resolved — booking amended with no change fee; Rs. 350 fare difference to be refunded to original card. "
            "Action items: Send updated confirmation email to customer (already done). "
            "Sentiment: Satisfied. "
            "Key details: Booking ETH-GOA-55821, Hotel Solora, deluxe room, refund Rs. 350, agent Divya, customer Sameer Khan."
        ),
        "flaw": "wrong_number",
        "what_wrong": "A flawed summary may state a different fare-difference amount or different dates.",
    },
    {
        "scenario": "Group booking inquiry — complex, multiple action items",
        "transcript": _t(
            ("Agent", "EasyTravel group desk, this is Pooja speaking."),
            ("Customer", "Hi Pooja, this is Manish from Lumen Edutech. We have a corporate group of 18 people travelling to Phuket for an offsite."),
            ("Agent", "Hello Manish, happy to assist. What dates are you considering?"),
            ("Customer", "Outbound Bangalore to Phuket on 5 November, return on 9 November."),
            ("Agent", "Got it. Any specific flight preferences? Direct or with one stop?"),
            ("Customer", "Direct if possible, but flexible if cost is high."),
            ("Agent", "Let me check availability... For 18 passengers direct, I am seeing 16 seats on the early morning EasyTravel flight and 2 seats on the partner airline."),
            ("Customer", "Hmm, can we split or shall we look at one-stop?"),
            ("Agent", "One-stop via Kuala Lumpur is fully available for the full group and saves about 12,000 rupees per ticket. Total fare difference for 18 would be around 2,16,000."),
            ("Customer", "That's significant. Let's go with one-stop. What about hotel?"),
            ("Agent", "For Phuket we have partner hotels with group rates. Two options: 4-star Resort Bayside at 6,800 per night per room, or 5-star Lumera Phuket at 11,500 per night."),
            ("Customer", "Bayside sounds reasonable. We need 9 twin-share rooms for 4 nights."),
            ("Agent", "9 rooms x 4 nights x 6,800 = 2,44,800 for accommodation, before taxes."),
            ("Customer", "Okay. Can you also arrange airport transfers and one team dinner?"),
            ("Agent", "Yes — coach transfer both ways at 12,000 total. Team dinner at the hotel beachside restaurant at 1,800 per person for the set menu."),
            ("Customer", "Add that. Can you send a formal quotation by tomorrow morning?"),
            ("Agent", "Absolutely. I will email a detailed quote to manish@lumen-edutech.example by 11 AM tomorrow."),
            ("Customer", "And what's the deposit structure?"),
            ("Agent", "30 percent deposit to confirm, balance 30 days before departure."),
            ("Customer", "Understood. Send across the quote and we will move forward."),
            ("Agent", "Done. Quote tomorrow by 11 AM, with deposit invoice attached."),
        ),
        "expected": (
            "Corporate group of 18 from Lumen Edutech for Bangalore-Phuket, 5-9 Nov. Agent Pooja "
            "recommended one-stop via KL (saves Rs. 2,16,000 total), Resort Bayside 4-star (Rs. 6,800/night, "
            "9 twin rooms x 4 nights = Rs. 2,44,800), coach transfers Rs. 12,000, beachside team dinner "
            "Rs. 1,800/person. Quote and deposit invoice (30%) to be emailed by 11 AM next day."
        ),
        "correct": (
            "Issue: Corporate group booking inquiry, 18 pax, Bangalore-Phuket, 5-9 Nov 2024. "
            "Resolution: Pending — quotation to be sent for approval. "
            "Action items: Agent Pooja to email detailed quote to manish@lumen-edutech.example by 11 AM next day, with deposit invoice (30% to confirm, balance 30 days before departure). "
            "Sentiment: Satisfied. "
            "Key details: 18 pax, one-stop via KL flights, Resort Bayside (9 twin x 4 nights x Rs. 6,800 = Rs. 2,44,800), coach Rs. 12,000, dinner Rs. 1,800/pax."
        ),
        "flaw": "missed_action_item",
        "what_wrong": "A flawed summary may drop the deposit invoice or the airport-transfer arrangement.",
    },
    {
        "scenario": "Flight rescheduling — medical emergency, empathetic handling",
        "transcript": _t(
            ("Agent", "EasyTravel, this is Aakash. How may I assist?"),
            ("Customer", "Hello Aakash, I am Reema Pillai. I need to reschedule my flight ET-6622 from Kolkata to London on the 25th. My father has had a heart attack."),
            ("Agent", "I am so sorry to hear that, Reema. I will do everything I can to help. May I have your booking reference?"),
            ("Customer", "PNR is KX9PL7."),
            ("Agent", "Found it. The current booking is 25 September Kolkata to London Heathrow. What new date are you considering?"),
            ("Customer", "I cannot say exactly. He just got admitted. Can I keep it open?"),
            ("Agent", "Yes. For medical emergencies we can convert the ticket into an open ticket valid for 6 months from the original date. No rebooking fee, only any future fare difference."),
            ("Customer", "Thank you so much, that is a relief."),
            ("Agent", "Do you have any documentation? An admission slip or doctor's note will help us record this as a medical waiver case."),
            ("Customer", "Yes I have the admission letter from the hospital."),
            ("Agent", "Please email it to medical-waivers@easytravel.example with subject 'KX9PL7 medical'. Within 24 hours of receipt I will email back the formal open-ticket confirmation."),
            ("Customer", "I will send it right after this call."),
            ("Agent", "I am also waiving the call-centre processing fee. You will not see any deductions."),
            ("Customer", "Thank you, that's very kind."),
            ("Agent", "Of course. When you are ready to fly, just reply to that email or call us with a new date and we will rebook subject to availability."),
            ("Customer", "Got it. Thanks Aakash, take care."),
            ("Agent", "You take care of yourself and your father. Wishing him a speedy recovery."),
        ),
        "expected": (
            "Customer Reema Pillai requested rescheduling of flight ET-6622 / PNR KX9PL7 (CCU-LHR, 25 Sep) "
            "due to her father's heart attack. Agent Aakash converted booking to an open ticket valid 6 months "
            "from original date, with no rebooking fee or processing fee. Customer to email admission letter "
            "to medical-waivers@easytravel.example for formal confirmation within 24 hours. Resolved."
        ),
        "correct": (
            "Issue: Flight ET-6622 / PNR KX9PL7 (CCU-LHR, 25 Sep) reschedule request due to father's medical emergency. "
            "Resolution: Resolved — converted to open ticket valid 6 months from original date, no rebooking or processing fee. "
            "Action items: Customer to email admission letter to medical-waivers@easytravel.example; agent to send formal open-ticket confirmation within 24 hours of receipt. "
            "Sentiment: Distressed but appreciative — overall positive interaction. "
            "Key details: PNR KX9PL7, ET-6622, CCU-LHR, agent Aakash, customer Reema Pillai."
        ),
        "flaw": "wrong_sentiment",
        "what_wrong": "A flawed summary may simply label the customer 'Satisfied' without conveying the emotional context, or label them 'Frustrated' which mischaracterizes the empathetic conversation.",
    },
    {
        "scenario": "Duplicate charge dispute — escalated to billing",
        "transcript": _t(
            ("Agent", "EasyTravel billing, you have Naveen on the line."),
            ("Customer", "Hi Naveen, I am Tanvi Bhatt. I see two charges of 9,200 rupees each on my credit card for the same booking. It should only be one."),
            ("Agent", "I am sorry about that, Tanvi. Please share the booking reference."),
            ("Customer", "It is ET-BNG-77291."),
            ("Agent", "One moment... I can see the booking. There is one payment authorisation that succeeded, but I also see a separate transaction with status 'failed' on our side."),
            ("Customer", "But my bank shows both as posted on my statement."),
            ("Agent", "That can happen when the bank holds the amount even after a failure. Could you read me the last 4 digits of the card used?"),
            ("Customer", "4471."),
            ("Agent", "Confirmed. I can see only one settled payment from our side. The duplicate is likely a stuck authorisation by your bank."),
            ("Customer", "What do I do? I need that money back."),
            ("Agent", "Two parallel actions. First, I will raise a dispute on our side with our payment gateway and provide you a written statement that only one charge was settled by EasyTravel."),
            ("Customer", "Okay."),
            ("Agent", "Second, you should contact your card issuer with that statement and they will reverse the second authorisation typically within 5 to 7 business days."),
            ("Customer", "Can you escalate this so it doesn't take ages?"),
            ("Agent", "Yes. I am escalating to our billing senior, Lakshmi Iyer, who will own this case end-to-end. Escalation ID BILL-30911."),
            ("Customer", "Will she contact me?"),
            ("Agent", "Lakshmi will email you the gateway statement within 48 hours so you can take it to your bank. Reference BILL-30911."),
            ("Customer", "Thanks. I'd appreciate that."),
            ("Agent", "Of course. I'll also send you a summary of this call by SMS."),
        ),
        "expected": (
            "Customer Tanvi Bhatt reported a duplicate Rs. 9,200 charge on booking ET-BNG-77291. "
            "Agent Naveen confirmed only one charge was settled on the EasyTravel side and identified "
            "the other as a stuck bank authorisation. Escalated to billing senior Lakshmi Iyer "
            "(BILL-30911); gateway statement to be emailed within 48 hours. Escalated."
        ),
        "correct": (
            "Issue: Duplicate Rs. 9,200 charge on booking ET-BNG-77291; one is a stuck bank authorisation, not a settled charge. "
            "Resolution: Escalated to billing senior Lakshmi Iyer under reference BILL-30911. "
            "Action items: Lakshmi to email gateway statement to customer within 48 hours; customer to use it with card issuer to reverse the stuck authorisation (typically 5-7 business days). "
            "Sentiment: Concerned but cooperative — Neutral/Satisfied by call end. "
            "Key details: Booking ET-BNG-77291, card ending 4471, escalation BILL-30911, agent Naveen, customer Tanvi Bhatt."
        ),
        "flaw": "wrong_resolution",
        "what_wrong": "A flawed summary may mark this as 'Resolved' even though it was escalated and is still pending the gateway statement.",
    },
    {
        "scenario": "Honeymoon package customization — happy customer, upsell",
        "transcript": _t(
            ("Agent", "Welcome to EasyTravel holidays, you have Ishita speaking."),
            ("Customer", "Hi Ishita, I am Karthik Reddy. My fiancée Anjali and I are looking at the Maldives honeymoon package."),
            ("Agent", "Congratulations! Which dates are you considering?"),
            ("Customer", "Travel from 18 December to 23 December. 5 nights."),
            ("Agent", "Wonderful. Our standard Maldives Honeymoon Bliss package is 1,15,000 per couple. Want me to walk you through what is included?"),
            ("Customer", "Yes please."),
            ("Agent", "It includes return flights Bangalore to Male, 5 nights at a water villa at Coral Pearl Resort, daily breakfast and dinner, one candle-lit beach dinner, a 60-minute couple spa, and seaplane transfers."),
            ("Customer", "Sounds nice. Any room upgrade options?"),
            ("Agent", "For 25,000 extra per couple we can upgrade to the Sunset Water Villa with a private plunge pool."),
            ("Customer", "Tempting. What about the spa duration?"),
            ("Agent", "Standard is 60 minutes per person. We can upgrade to 90 minutes for 6,500 per couple."),
            ("Customer", "Let's do both — sunset villa and 90-minute spa."),
            ("Agent", "Excellent. New total: 1,15,000 + 25,000 + 6,500 = 1,46,500 per couple. Is that within your budget?"),
            ("Customer", "Yes that's good. Are there any complimentary perks for honeymooners?"),
            ("Agent", "Yes — complimentary cake on arrival, room decoration with flowers, and a sunset cruise on day 3."),
            ("Customer", "Wow, lovely!"),
            ("Agent", "Shall I go ahead and book? I will need a 30 percent advance to confirm, balance 45 days before travel."),
            ("Customer", "Yes please. I will make the advance payment today."),
            ("Agent", "Perfect. Booking reference HON-MV-22441. Advance is 43,950, balance 1,02,550 due by 3 November. Confirmation invoice has been emailed."),
            ("Customer", "Thanks Ishita. We are really excited!"),
            ("Agent", "I am thrilled for you both. Wishing Karthik and Anjali a magical honeymoon."),
        ),
        "expected": (
            "Customer Karthik Reddy booked the Maldives Honeymoon package (Coral Pearl Resort, 5 nights, "
            "18-23 Dec) for couple. Upgraded to Sunset Water Villa (+Rs. 25,000) and 90-min couple spa "
            "(+Rs. 6,500). Total Rs. 1,46,500. Booking ref HON-MV-22441; advance Rs. 43,950 today, "
            "balance Rs. 1,02,550 due by 3 Nov. Resolved."
        ),
        "correct": (
            "Issue: Customisation and booking of Maldives Honeymoon package for couple, 18-23 Dec 2024. "
            "Resolution: Resolved — booking HON-MV-22441 created with Sunset Water Villa upgrade and 90-minute couple spa. "
            "Action items: Customer to pay 30% advance (Rs. 43,950) today; balance Rs. 1,02,550 by 3 November. "
            "Sentiment: Satisfied (excited). "
            "Key details: Booking HON-MV-22441, Coral Pearl Resort, total Rs. 1,46,500, agent Ishita, customer Karthik Reddy (fiancée Anjali)."
        ),
        "flaw": "fabricated_detail",
        "what_wrong": "A flawed summary may invent an extra inclusion (e.g., 'free airport lounge access' or 'free GoPro rental') that was not mentioned in the transcript.",
    },
]


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
ALT_FILL = PatternFill("solid", fgColor="F0F4FA")


def _write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_row(ws, row_idx, ncols, alt):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sample Transcripts"
    _write_header(ws1, ["Scenario", "Full_Transcript", "Expected_Summary"])
    for i, item in enumerate(TRANSCRIPTS, start=2):
        ws1.cell(row=i, column=1, value=item["scenario"])
        ws1.cell(row=i, column=2, value=item["transcript"])
        ws1.cell(row=i, column=3, value=item["expected"])
        ws1.row_dimensions[i].height = 200
        _style_row(ws1, i, 3, alt=(i % 2 == 0))
    _set_col_widths(ws1, [45, 80, 60])

    ws2 = wb.create_sheet("Evaluation Sheet")
    _write_header(ws2, ["Scenario", "AI_Generated_Summary", "Your_Evaluation", "Notes"])
    for i, item in enumerate(TRANSCRIPTS, start=2):
        ws2.cell(row=i, column=1, value=item["scenario"])
        ws2.cell(row=i, column=2, value="")
        ws2.cell(row=i, column=3, value="")
        ws2.cell(row=i, column=4, value="")
        _style_row(ws2, i, 4, alt=(i % 2 == 0))
    _set_col_widths(ws2, [45, 60, 25, 40])

    ws3 = wb.create_sheet("Answer Key")
    _write_header(ws3, ["Scenario", "Correct_Summary", "Known_Flaw_Type", "What_Was_Wrong"])
    for i, item in enumerate(TRANSCRIPTS, start=2):
        ws3.cell(row=i, column=1, value=item["scenario"])
        ws3.cell(row=i, column=2, value=item["correct"])
        ws3.cell(row=i, column=3, value=item["flaw"])
        ws3.cell(row=i, column=4, value=item["what_wrong"])
        _style_row(ws3, i, 4, alt=(i % 2 == 0))
    _set_col_widths(ws3, [40, 70, 22, 60])
    ws3.sheet_state = "hidden"

    wb.save(OUTPUT_FILE)
    print(f"  generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
